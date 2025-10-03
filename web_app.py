import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import os

st.title("Excel Merger: Quantities & Dates with Product Info")

# --- Load prod_line.xlsx automatically ---
prod_line_path = os.path.join(os.path.dirname(__file__), "prod_line.xlsx")
if not os.path.exists(prod_line_path):
    st.error("prod_line.xlsx not found in app folder!")
    st.stop()

prod_line = pd.read_excel(prod_line_path)
prod_line.columns = prod_line.columns.str.strip()  # normalize

required_prod_cols = ['bar_code','export_code','prod_description','category','price']
for col in required_prod_cols:
    if col not in prod_line.columns:
        st.error(f"Column '{col}' missing in prod_line.xlsx")
        st.stop()

# unify codes as strings with no leading zeros/whitespace
prod_line['bar_code'] = prod_line['bar_code'].astype(str).str.strip().str.lstrip("0")

# --- Upload multiple Excel files ---
uploaded_files = st.file_uploader(
    "Upload Excel files", type=["xlsx", "xls"], accept_multiple_files=True
)

if uploaded_files:
    df_list = []
    dates_list = []
    highlight_dict = {}  # track kodes that need red-fill per location
    per_file_sheets = {}  # store per-file sheets for appending later

    for uploaded_file in uploaded_files:
        location = uploaded_file.name.split(".")[0]
        df = pd.read_excel(uploaded_file)
        df.columns = df.columns.str.strip()

        if 'კოდი' not in df.columns or 'რაოდენობა' not in df.columns:
            st.error(f"Columns missing in {uploaded_file.name}")
            st.stop()

        if 'თვე' not in df.columns:
            df['თვე'] = None
        if 'წელი' not in df.columns:
            df['წელი'] = None

        # unify code format
        df['კოდი'] = df['კოდი'].astype(str).str.strip().str.lstrip("0")

        # --- Quantities for merged sheet ---
        df_qty = df.groupby('კოდი', as_index=False)['რაოდენობა'].sum().rename(columns={'რაოდენობა': location})
        df_list.append(df_qty)

        # --- Dates for merged sheet ---
        df_dates = df.groupby('კოდი', as_index=False).agg({'თვე':'first', 'წელი':'first'})
        df_dates['თვე_num'] = pd.to_numeric(df_dates['თვე'], errors='coerce')
        df_dates['წელი_num'] = pd.to_numeric(df_dates['წელი'], errors='coerce')

        # Collect kodes that need highlighting
        kodes_to_highlight = df_dates[
            (df_dates['თვე_num'].isna()) |
            (df_dates['წელი_num'].isna()) |
            ((df_dates['წელი_num'] == 2025) & (df_dates['თვე_num'] < 10))
        ]['კოდი'].astype(str).tolist()
        highlight_dict[location] = kodes_to_highlight

        # Format month/year
        def format_date(row):
            if pd.notnull(row['თვე_num']) and pd.notnull(row['წელი_num']):
                return f"{int(row['თვე_num']):02d}/{int(row['წელი_num'])}"
            else:
                return ""
        df_dates[location] = df_dates.apply(format_date, axis=1)
        df_dates = df_dates[['კოდი', location]]
        dates_list.append(df_dates)

        # --- Prepare per-file sheet ---
        df_file = df.copy()
        df_file = pd.merge(df_file, prod_line, left_on='კოდი', right_on='bar_code', how='left')
        df_file.drop(columns=['bar_code'], inplace=True)
        # Add index column at the very left
        df_file.insert(0, 'Index', range(1, len(df_file)+1))
        per_file_sheets[location] = df_file

    # --- Merge Quantities ---
    final_qty = df_list[0]
    for df in df_list[1:]:
        final_qty = pd.merge(final_qty, df, on='კოდი', how='outer')
    final_qty = final_qty.fillna(0)

    # --- Merge Dates ---
    final_dates = dates_list[0]
    for df in dates_list[1:]:
        final_dates = pd.merge(final_dates, df, on='კოდი', how='outer')
    final_dates = final_dates.fillna("")

    # --- Merge with product info ---
    final_qty = pd.merge(final_qty, prod_line, left_on='კოდი', right_on='bar_code', how='left')
    final_dates = pd.merge(final_dates, prod_line, left_on='კოდი', right_on='bar_code', how='left')
    final_qty.drop(columns=['bar_code'], inplace=True)
    final_dates.drop(columns=['bar_code'], inplace=True)

    prod_cols = ['export_code','prod_description','category','price']
    for col in prod_cols:
        if col not in final_qty.columns:
            final_qty[col] = None
        if col not in final_dates.columns:
            final_dates[col] = None

    # --- Reorder columns ---
    qty_cols = ['კოდი'] + prod_cols + [c for c in final_qty.columns if c not in ['კოდი'] + prod_cols]
    final_qty = final_qty[qty_cols]
    date_cols = ['კოდი'] + prod_cols + [c for c in final_dates.columns if c not in ['კოდი'] + prod_cols]
    final_dates = final_dates[date_cols]

    # --- Show tables ---
    st.subheader("Quantities with Product Info")
    st.dataframe(final_qty)
    st.subheader("Matched Dates with Product Info")
    st.dataframe(final_dates)

    # --- Save to Excel ---
    temp_output = BytesIO()
    with pd.ExcelWriter(temp_output, engine='openpyxl') as writer:
        # main merged sheets
        final_qty.to_excel(writer, sheet_name='Quantities', index=False)
        final_dates.to_excel(writer, sheet_name='Matched Dates', index=False)
        # uploaded files with product info
        for sheet_name, df_file in per_file_sheets.items():
            df_file.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    temp_output.seek(0)
    wb = load_workbook(temp_output)

    # --- Formatting ---
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    for ws in wb.worksheets:
        for col in ws.columns:
            column_letter = col[0].column_letter
            max_length = 0
            for cell in col:
                cell.border = thin_border
                cell.alignment = align_center
                if cell.row == 1:
                    cell.font = header_font
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2

    # --- Highlight missing/early dates in main sheets ---
    for sheet_name in ['Quantities', 'Matched Dates']:
        ws = wb[sheet_name]
        for loc, highlight_kodes in highlight_dict.items():
            # find column index of location
            col_idx = None
            for idx, cell in enumerate(ws[1], start=1):
                if cell.value == loc:
                    col_idx = idx
                    break
            if col_idx is None:
                continue
            for row_idx in range(2, ws.max_row + 1):
                kode = ws.cell(row=row_idx, column=1).value
                if str(kode) in highlight_kodes:
                    ws.cell(row=row_idx, column=col_idx).fill = red_fill

    # --- Highlight missing/early dates in per-file sheets ---
    for sheet_name, df_file in per_file_sheets.items():
        ws = wb[sheet_name[:31]]
        # find column index of 'თვე' or 'წელი' for highlighting
        col_idx = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value in ['თვე','წელი']:
                col_idx = idx
                break
        if col_idx is None:
            continue
        for row_idx in range(2, ws.max_row + 1):
            try:
                month_cell = ws.cell(row=row_idx, column=ws[1].index(next(c for c in ws[1] if c.value=='თვე'))+1).value
                year_cell = ws.cell(row=row_idx, column=ws[1].index(next(c for c in ws[1] if c.value=='წელი'))+1).value
                month = pd.to_numeric(month_cell, errors='coerce')
                year = pd.to_numeric(year_cell, errors='coerce')
                if pd.isna(month) or pd.isna(year) or (year==2025 and month<10):
                    for cell in ws[row_idx]:
                        cell.fill = red_fill
            except:
                continue

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    st.download_button(
        label="Download Combined Excel with Product Info",
        data=final_output,
        file_name="combined_locations_with_prod_info.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
